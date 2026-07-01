"""Export service — structured documents to JSON, CSV, Excel, text reports, ZIP.

Usage::

    from services.export_service import ExportService, ExportFormat

    svc = ExportService()

    # Single document.
    result = svc.export_single(doc, ExportFormat.JSON)

    # Batch export.
    results = svc.export_batch(all_docs, ExportFormat.CSV)

    # Metadata-only (JSON).
    result = svc.export_metadata(all_docs)

    # ZIP bundle containing JSON + original files.
    result = svc.export_zip(all_docs, include_originals=True)
"""

from __future__ import annotations

import csv
import io
import json
import time
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any

from config import settings
from utils.file_utils import ensure_dir
from utils.logger import get_logger

logger = get_logger(__name__)

# openpyxl is optional — Excel export requires it.
_OPENPYXL_AVAILABLE = False
try:
    import openpyxl as _xl
    from openpyxl.styles import Font, Alignment, PatternFill

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _xl = None  # type: ignore[assignment]


# =========================================================================
# Enums & data models
# =========================================================================


class ExportFormat(str, Enum):
    """Supported export output formats."""

    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    REPORT = "report"
    ZIP = "zip"


@dataclass
class ExportDocument:
    """Normalised document representation for all export formats.

    Created via the :meth:`from_record` factory which accepts either a
    ``services.document_service.DocumentRecord`` (in-memory) or a
    ``database.models.Document`` (SQLAlchemy ORM).
    """

    id: str
    filename: str
    file_path: str
    file_size: int
    mime_type: str
    file_hash: str
    status: str
    created_at: float
    updated_at: float
    extracted_text: str | None = None
    structured_json: dict[str, Any] | None = None
    confidence_score: float = 0.0
    error_message: str | None = None
    processing_time: float = 0.0

    @classmethod
    def from_record(cls, record: Any) -> ExportDocument:
        """Build from a ``DocumentRecord`` or SQLAlchemy ``Document``."""
        if _is_dataclass(record):
            return cls(
                id=record.id,
                filename=record.filename,
                file_path=str(record.file_path),
                file_size=record.file_size,
                mime_type=record.mime_type,
                file_hash=record.file_hash,
                status=_str_status(record.status),
                created_at=record.created_at,
                updated_at=record.updated_at,
                extracted_text=getattr(record, "extracted_text", None),
                structured_json=getattr(record, "structured_json", None),
                confidence_score=getattr(record, "confidence_score", 0.0),
                error_message=getattr(record, "error_message", None),
                processing_time=getattr(record, "processing_time", 0.0),
            )
        # SQLAlchemy ORM model.
        try:
            from database.models import Document as ORMDocument

            if isinstance(record, ORMDocument):
                return cls(
                    id=record.id,
                    filename=record.filename,
                    file_path=record.file_path,
                    file_size=record.file_size,
                    mime_type=record.mime_type,
                    file_hash=record.file_hash,
                    status=record.status,
                    created_at=record.created_at,
                    updated_at=record.updated_at,
                    extracted_text=record.extracted_text,
                    structured_json=_parse_json(record.structured_json),
                    confidence_score=record.confidence_score,
                    error_message=record.error_message,
                    processing_time=record.processing_time,
                )
        except ImportError:
            pass
        raise TypeError(f"Cannot convert {type(record).__name__} to ExportDocument")

    def to_dict(self, include_text: bool = True) -> dict[str, Any]:
        """Serialise to a plain dict for JSON/CSV export.

        Args:
            include_text: When ``False``, omit ``extracted_text`` and
                ``structured_json`` (metadata-only export).
        """
        d: dict[str, Any] = {
            "id": self.id,
            "filename": self.filename,
            "file_path": self.file_path,
            "file_size": self.file_size,
            "mime_type": self.mime_type,
            "file_hash": self.file_hash,
            "status": self.status,
            "created_at": self.created_at,
            "updated_at": self.updated_at,
            "confidence_score": self.confidence_score,
            "error_message": self.error_message,
            "processing_time": self.processing_time,
        }
        if include_text:
            d["extracted_text"] = self.extracted_text
            d["structured_json"] = self.structured_json
        return d

    def to_metadata_dict(self) -> dict[str, Any]:
        """Dict with metadata fields only (no extracted content)."""
        return self.to_dict(include_text=False)


@dataclass
class ExportResult:
    """Outcome of a single export operation."""

    path: Path
    format: ExportFormat
    count: int
    size_bytes: int
    timestamp: float = 0.0

    def __post_init__(self) -> None:
        if self.timestamp == 0.0:
            object.__setattr__(self, "timestamp", time.time())


class ExportError(Exception):
    """Raised when an export operation fails."""


# =========================================================================
# ExportService
# =========================================================================


class ExportService:
    """Export documents in JSON, CSV, Excel, text-report, and ZIP formats.

    All output files are written to ``{DATA_DIR}/exports/`` by default.
    Timestamped filenames are auto-generated when *path* is ``None``.
    """

    def __init__(self, output_dir: Path | None = None) -> None:
        self._output_dir = Path(output_dir or settings.DATA_DIR / "exports")
        self._output_dir.mkdir(parents=True, exist_ok=True)

    # ── Convenience ─────────────────────────────────────────────────────

    def export_single(
        self,
        document: Any,
        fmt: ExportFormat | str,
        path: Path | None = None,
    ) -> ExportResult:
        """Export a single document."""
        return self.export_batch([document], fmt, path=path)

    def export_batch(
        self,
        documents: Sequence[Any],
        fmt: ExportFormat | str,
        path: Path | None = None,
    ) -> ExportResult:
        """Export multiple documents in the specified format."""
        fmt = ExportFormat(fmt) if isinstance(fmt, str) else fmt
        docs = self._normalise(documents)
        path = path or self._default_path(fmt)

        exporters = {
            ExportFormat.JSON: self._export_json,
            ExportFormat.CSV: self._export_csv,
            ExportFormat.EXCEL: self._export_excel,
            ExportFormat.REPORT: self._export_report,
        }

        exporter = exporters.get(fmt)
        if exporter is None:
            raise ExportError(f"Unsupported format: {fmt}")

        ensure_dir(path.parent)
        exporter(docs, path)

        size = path.stat().st_size
        logger.info("Exported %d doc(s) → %s (%s, %d bytes)", len(docs), path.name, fmt.value, size)
        return ExportResult(path=path, format=fmt, count=len(docs), size_bytes=size)

    def export_metadata(
        self,
        documents: Sequence[Any],
        path: Path | None = None,
    ) -> ExportResult:
        """Export metadata-only (no extracted text) as JSON."""
        docs = self._normalise(documents)
        path = path or self._default_path(ExportFormat.JSON, suffix="metadata")
        ensure_dir(path.parent)

        data = [d.to_metadata_dict() for d in docs]
        path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
        size = path.stat().st_size
        logger.info("Exported metadata for %d doc(s) → %s (%d bytes)", len(docs), path.name, size)
        return ExportResult(path=path, format=ExportFormat.JSON, count=len(docs), size_bytes=size)

    def export_zip(
        self,
        documents: Sequence[Any],
        path: Path | None = None,
        include_originals: bool = False,
    ) -> ExportResult:
        """Bundle multiple export formats and optionally original files.

        The ZIP always contains:
        - ``documents.json`` — full document data
        - ``metadata.json`` — metadata-only
        - ``documents.csv`` — metadata as CSV
        - ``report.txt`` — human-readable summary

        When *include_originals* is ``True``, original uploaded files are
        included (if they still exist on disk).
        """
        docs = self._normalise(documents)
        path = path or self._default_path(ExportFormat.ZIP)
        ensure_dir(path.parent)

        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
            # Full JSON
            json_data = json.dumps(
                [d.to_dict() for d in docs], indent=2, default=str
            ).encode("utf-8")
            zf.writestr("documents.json", json_data)

            # Metadata JSON
            meta_data = json.dumps(
                [d.to_metadata_dict() for d in docs], indent=2, default=str
            ).encode("utf-8")
            zf.writestr("metadata.json", meta_data)

            # CSV
            csv_buffer = io.StringIO()
            self._write_csv(docs, csv_buffer)
            zf.writestr("documents.csv", csv_buffer.getvalue().encode("utf-8"))

            # Report
            report_text = self._build_report(docs)
            zf.writestr("report.txt", report_text.encode("utf-8"))

            # Original files (optional).
            if include_originals:
                for d in docs:
                    src = Path(d.file_path)
                    if src.is_file():
                        safe_name = Path(d.filename).name  # strip directory components
                        zf.write(src, f"originals/{safe_name}")
                    else:
                        logger.warning("Original file not found: %s", src)

        size = path.stat().st_size
        logger.info("ZIP export (%d docs) → %s (%d bytes)", len(docs), path.name, size)
        return ExportResult(path=path, format=ExportFormat.ZIP, count=len(docs), size_bytes=size)

    # ── Internal exporters ──────────────────────────────────────────────

    @staticmethod
    def _export_json(docs: list[ExportDocument], path: Path) -> None:
        data = [d.to_dict() for d in docs]
        path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")

    @staticmethod
    def _export_csv(docs: list[ExportDocument], path: Path) -> None:
        with path.open("w", newline="", encoding="utf-8") as fh:
            ExportService._write_csv(docs, fh)

    @staticmethod
    def _write_csv(docs: list[ExportDocument], fh: Any) -> None:
        if not docs:
            return
        fieldnames = list(docs[0].to_dict().keys())
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for d in docs:
            row = d.to_dict()
            row["structured_json"] = json.dumps(row["structured_json"], default=str) if row["structured_json"] else ""
            writer.writerow(row)

    @staticmethod
    def _export_excel(docs: list[ExportDocument], path: Path) -> None:
        if not _OPENPYXL_AVAILABLE:
            raise ExportError(
                "Excel export requires openpyxl. Install it with: pip install openpyxl"
            )
        wb = _xl.Workbook()
        # Sheet 1 — full data.
        ws = wb.active
        ws.title = "Documents"
        ExportService._write_excel_sheet(ws, docs, include_text=True)

        # Sheet 2 — metadata only.
        ws2 = wb.create_sheet("Metadata")
        ExportService._write_excel_sheet(ws2, docs, include_text=False)

        wb.save(str(path))

    @staticmethod
    def _write_excel_sheet(
        ws: Any,
        docs: list[ExportDocument],
        include_text: bool,
    ) -> None:
        if not docs:
            return
        rows = [d.to_dict(include_text=include_text) for d in docs]
        headers = list(rows[0].keys())
        # Header row with styling.
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                val = row_data[key]
                if key == "structured_json" and val is not None:
                    val = json.dumps(val, default=str)
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width (capped).
        for i, header in enumerate(headers, 1):
            max_len = max(
                len(str(header)),
                max((len(str(row[key])) for row in rows if row.get(key)), default=0),
            )
            ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{i}"].width = min(max_len + 3, 60)

    @staticmethod
    def _export_report(docs: list[ExportDocument], path: Path) -> None:
        report = ExportService._build_report(docs)
        path.write_text(report, encoding="utf-8")

    @staticmethod
    def _build_report(docs: list[ExportDocument]) -> str:
        lines: list[str] = []
        sep = "=" * 78
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        lines.append(sep)
        lines.append(" LOCAL AI STRUCTIFY — EXPORT REPORT")
        lines.append(f" Generated: {now}")
        lines.append(f" Documents: {len(docs)}")
        lines.append(sep)
        lines.append("")

        for i, d in enumerate(docs, 1):
            lines.append(f"Document {i}: {d.filename}")
            lines.append("-" * 78)
            lines.append(f"  ID:              {d.id}")
            lines.append(f"  Type:            {d.mime_type}")
            lines.append(f"  Status:          {d.status}")
            lines.append(f"  Size:            {_human_size(d.file_size)}")
            lines.append(f"  Created:         {_format_ts(d.created_at)}")
            lines.append(f"  Updated:         {_format_ts(d.updated_at)}")
            lines.append(f"  Confidence:      {d.confidence_score:.2f}")
            lines.append(f"  Processing time: {d.processing_time:.1f}s")
            if d.extracted_text:
                lines.append(f"  Text length:     {len(d.extracted_text):,} chars")
            if d.structured_json:
                lines.append(f"  Schema type:     {d.structured_json.get('type', 'N/A')}")
            if d.error_message:
                lines.append(f"  Error:           {d.error_message}")
            lines.append("")
            if d.extracted_text:
                preview = d.extracted_text[:200].replace("\n", " ")
                lines.append(f"  Preview (first 200 chars):")
                lines.append(f"    {preview}")
            lines.append("")
            lines.append(sep)
            lines.append("")

        return "\n".join(lines)

    # ── Helpers ─────────────────────────────────────────────────────────

    def _normalise(self, documents: Sequence[Any]) -> list[ExportDocument]:
        if not documents:
            return []
        if all(isinstance(d, ExportDocument) for d in documents):
            return list(documents)
        return [ExportDocument.from_record(d) for d in documents]

    def _default_path(self, fmt: ExportFormat, suffix: str = "") -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name_parts = [f"export_{timestamp}"]
        if suffix:
            name_parts.append(suffix)
        ext = _FORMAT_EXTENSIONS.get(fmt, ".dat")
        filename = "_".join(name_parts) + ext
        return self._output_dir / filename


_FORMAT_EXTENSIONS: dict[ExportFormat, str] = {
    ExportFormat.JSON: ".json",
    ExportFormat.CSV: ".csv",
    ExportFormat.EXCEL: ".xlsx",
    ExportFormat.REPORT: ".txt",
    ExportFormat.ZIP: ".zip",
}


# =========================================================================
# Internal helpers
# =========================================================================


def _is_dataclass(obj: Any) -> bool:
    return hasattr(obj, "__dataclass_fields__")


def _str_status(status: Any) -> str:
    """Convert enum/string status to plain string."""
    if isinstance(status, str):
        return status
    try:
        return status.value
    except AttributeError:
        return str(status)


def _parse_json(value: str | None) -> dict[str, Any] | None:
    if value is None:
        return None
    try:
        return json.loads(value)
    except (json.JSONDecodeError, TypeError):
        return None


def _human_size(bytes_: int) -> str:
    if bytes_ < 1024:
        return f"{bytes_} B"
    kb = bytes_ / 1024
    if kb < 1024:
        return f"{kb:.1f} KB"
    mb = kb / 1024
    if mb < 1024:
        return f"{mb:.1f} MB"
    gb = mb / 1024
    return f"{gb:.2f} GB"


def _format_ts(ts: float) -> str:
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
