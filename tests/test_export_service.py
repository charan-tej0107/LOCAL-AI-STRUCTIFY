"""Unit tests for Module 15: Export (services.export_service).

Uses temporary directories for all export output to isolate runs.
"""

from __future__ import annotations

import csv
import io
import json
import os
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pytest

from services.export_service import (
    ExportService,
    ExportDocument,
    ExportResult,
    ExportFormat,
    ExportError,
    _OPENPYXL_AVAILABLE,
    _human_size,
    _format_ts,
)


# =========================================================================
# Sample data
# =========================================================================


SAMPLE_DOCS = [
    ExportDocument(
        id="doc_001",
        filename="quarterly_report.pdf",
        file_path="/tmp/report.pdf",
        file_size=204800,
        mime_type="application/pdf",
        file_hash="abc123",
        status="processed",
        created_at=1000.0,
        updated_at=1100.0,
        extracted_text="Revenue grew 20% this quarter.",
        structured_json={"type": "report", "year": 2025, "amount": 50000},
        confidence_score=0.95,
        error_message=None,
        processing_time=15.2,
    ),
    ExportDocument(
        id="doc_002",
        filename="notes.txt",
        file_path="/tmp/notes.txt",
        file_size=512,
        mime_type="text/plain",
        file_hash="def456",
        status="uploaded",
        created_at=2000.0,
        updated_at=2100.0,
        extracted_text="Meeting notes about project Alpha.",
        structured_json=None,
        confidence_score=0.0,
        error_message="",
        processing_time=0.0,
    ),
    ExportDocument(
        id="doc_003",
        filename="héllo_世界.txt",
        file_path="/tmp/unicode.txt",
        file_size=256,
        mime_type="text/plain",
        file_hash="ghi789",
        status="failed",
        created_at=3000.0,
        updated_at=3100.0,
        extracted_text=None,
        structured_json=None,
        confidence_score=0.0,
        error_message="OCR failed: low confidence",
        processing_time=5.0,
    ),
]


# =========================================================================
# Fixtures
# =========================================================================


@pytest.fixture
def output_dir(tmp_path: Path) -> Path:
    return tmp_path / "exports"


@pytest.fixture
def export_svc(output_dir: Path) -> ExportService:
    return ExportService(output_dir=output_dir)


# =========================================================================
# ExportDocument model
# =========================================================================


class TestExportDocument:
    def test_to_dict_full(self) -> None:
        d = SAMPLE_DOCS[0].to_dict(include_text=True)
        assert d["id"] == "doc_001"
        assert d["extracted_text"] == "Revenue grew 20% this quarter."
        assert d["structured_json"] == {"type": "report", "year": 2025, "amount": 50000}

    def test_to_dict_metadata_only(self) -> None:
        d = SAMPLE_DOCS[0].to_dict(include_text=False)
        assert d["id"] == "doc_001"
        assert "extracted_text" not in d
        assert "structured_json" not in d

    def test_to_metadata_dict(self) -> None:
        d = SAMPLE_DOCS[0].to_metadata_dict()
        assert d["id"] == "doc_001"
        assert "extracted_text" not in d

    def test_from_record_dataclass(self) -> None:
        @dataclass
        class FakeRecord:
            id: str = "doc_001"
            filename: str = "test.txt"
            file_path: str = "/tmp/test.txt"
            file_size: int = 100
            mime_type: str = "text/plain"
            file_hash: str = "h1"
            status: str = "uploaded"
            created_at: float = 0.0
            updated_at: float = 0.0
            extracted_text: str | None = "hello"
            structured_json: dict | None = None
            confidence_score: float = 0.9
            error_message: str | None = None
            processing_time: float = 1.0

        doc = ExportDocument.from_record(FakeRecord())
        assert doc.id == "doc_001"
        assert doc.extracted_text == "hello"

    def test_from_record_invalid_type(self) -> None:
        with pytest.raises(TypeError, match="Cannot convert"):
            ExportDocument.from_record(42)  # type: ignore[arg-type]


# =========================================================================
# ExportService — JSON
# =========================================================================


class TestJsonExport:
    def test_single(self, export_svc: ExportService) -> None:
        result = export_svc.export_single(SAMPLE_DOCS[0], ExportFormat.JSON)
        assert result.format == ExportFormat.JSON
        assert result.count == 1
        assert result.path.is_file()
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert len(data) == 1
        assert data[0]["id"] == "doc_001"

    def test_batch(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch(SAMPLE_DOCS, ExportFormat.JSON)
        assert result.count == 3
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert len(data) == 3

    def test_empty(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([], ExportFormat.JSON)
        assert result.count == 0
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert data == []

    def test_custom_path(self, export_svc: ExportService, tmp_path: Path) -> None:
        custom = tmp_path / "custom.json"
        result = export_svc.export_batch(SAMPLE_DOCS, ExportFormat.JSON, path=custom)
        assert result.path == custom
        assert custom.is_file()

    def test_content_verification(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([SAMPLE_DOCS[0]], ExportFormat.JSON)
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert data[0]["filename"] == "quarterly_report.pdf"
        assert data[0]["structured_json"]["type"] == "report"
        assert data[0]["confidence_score"] == 0.95

    def test_string_format(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch(SAMPLE_DOCS, "json")
        assert result.format == ExportFormat.JSON


# =========================================================================
# ExportService — CSV
# =========================================================================


class TestCsvExport:
    def test_single(self, export_svc: ExportService) -> None:
        result = export_svc.export_single(SAMPLE_DOCS[0], ExportFormat.CSV)
        assert result.count == 1
        content = result.path.read_text(encoding="utf-8")
        assert "doc_001" in content
        assert "quarterly_report.pdf" in content

    def test_batch(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch(SAMPLE_DOCS, ExportFormat.CSV)
        assert result.count == 3
        reader = csv.DictReader(result.path.read_text(encoding="utf-8").splitlines())
        rows = list(reader)
        assert len(rows) == 3
        assert rows[0]["id"] == "doc_001"

    def test_header_present(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch(SAMPLE_DOCS[:1], ExportFormat.CSV)
        lines = result.path.read_text(encoding="utf-8").splitlines()
        assert lines[0] == (
            "id,filename,file_path,file_size,mime_type,file_hash,status,"
            "created_at,updated_at,confidence_score,error_message,"
            "processing_time,extracted_text,structured_json"
        )

    def test_empty(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([], ExportFormat.CSV)
        assert result.count == 0
        assert result.path.is_file()

    def test_structured_json_in_csv(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([SAMPLE_DOCS[0]], ExportFormat.CSV)
        content = result.path.read_text(encoding="utf-8")
        assert "report" in content
        assert "type" in content
        assert "2025" in content


# =========================================================================
# ExportService — Excel
# =========================================================================


class TestExcelExport:
    def test_requires_openpyxl(self, export_svc: ExportService) -> None:
        if not _OPENPYXL_AVAILABLE:
            with pytest.raises(ExportError, match="openpyxl"):
                export_svc.export_batch(SAMPLE_DOCS, ExportFormat.EXCEL)

    def test_batch_export(self, export_svc: ExportService) -> None:
        if not _OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
        result = export_svc.export_batch(SAMPLE_DOCS, ExportFormat.EXCEL)
        assert result.count == 3
        assert result.path.suffix == ".xlsx"
        assert result.path.is_file()

        import openpyxl
        wb = openpyxl.load_workbook(result.path)
        assert "Documents" in wb.sheetnames
        assert "Metadata" in wb.sheetnames
        ws = wb["Documents"]
        assert ws.cell(1, 1).value == "id"
        assert ws.cell(2, 1).value == "doc_001"

    def test_empty(self, export_svc: ExportService) -> None:
        if not _OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
        result = export_svc.export_batch([], ExportFormat.EXCEL)
        assert result.count == 0
        assert result.path.is_file()


# =========================================================================
# ExportService — Report
# =========================================================================


class TestReportExport:
    def test_contains_headers(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch(SAMPLE_DOCS, ExportFormat.REPORT)
        content = result.path.read_text(encoding="utf-8")
        assert "LOCAL AI STRUCTIFY — EXPORT REPORT" in content
        assert "Documents: 3" in content

    def test_contains_doc_info(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([SAMPLE_DOCS[0]], ExportFormat.REPORT)
        content = result.path.read_text(encoding="utf-8")
        assert "quarterly_report.pdf" in content
        assert "application/pdf" in content
        assert "processed" in content
        assert "0.95" in content

    def test_includes_text_preview(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([SAMPLE_DOCS[0]], ExportFormat.REPORT)
        content = result.path.read_text(encoding="utf-8")
        assert "Revenue grew 20%" in content

    def test_handles_none_text(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([SAMPLE_DOCS[2]], ExportFormat.REPORT)
        content = result.path.read_text(encoding="utf-8")
        assert "OCR failed:" in content
        assert "héllo_世界.txt" in content

    def test_empty(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([], ExportFormat.REPORT)
        content = result.path.read_text(encoding="utf-8")
        assert "Documents: 0" in content


# =========================================================================
# ExportService — Metadata export
# =========================================================================


class TestMetadataExport:
    def test_metadata_export(self, export_svc: ExportService) -> None:
        result = export_svc.export_metadata(SAMPLE_DOCS)
        assert result.count == 3
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert len(data) == 3
        # No extracted_text or structured_json in metadata export
        assert "extracted_text" not in data[0]
        assert "structured_json" not in data[0]

    def test_metadata_has_id(self, export_svc: ExportService) -> None:
        result = export_svc.export_metadata([SAMPLE_DOCS[0]])
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert data[0]["id"] == "doc_001"
        assert data[0]["filename"] == "quarterly_report.pdf"

    def test_empty(self, export_svc: ExportService) -> None:
        result = export_svc.export_metadata([])
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert data == []


# =========================================================================
# ExportService — ZIP
# =========================================================================


class TestZipExport:
    def test_zip_contents(self, export_svc: ExportService) -> None:
        result = export_svc.export_zip(SAMPLE_DOCS)
        assert result.format == ExportFormat.ZIP
        assert result.count == 3
        with zipfile.ZipFile(result.path) as zf:
            names = zf.namelist()
            assert "documents.json" in names
            assert "metadata.json" in names
            assert "documents.csv" in names
            assert "report.txt" in names

    def test_zip_json_content(self, export_svc: ExportService) -> None:
        result = export_svc.export_zip(SAMPLE_DOCS)
        with zipfile.ZipFile(result.path) as zf:
            data = json.loads(zf.read("documents.json"))
            assert len(data) == 3
            assert data[0]["id"] == "doc_001"

    def test_zip_metadata_content(self, export_svc: ExportService) -> None:
        result = export_svc.export_zip(SAMPLE_DOCS)
        with zipfile.ZipFile(result.path) as zf:
            data = json.loads(zf.read("metadata.json"))
            assert len(data) == 3
            assert "extracted_text" not in data[0]

    def test_zip_with_originals(self, export_svc: ExportService, tmp_path: Path) -> None:
        # Create real files for originals
        doc1 = tmp_path / "real_report.pdf"
        doc1.write_text("PDF content", encoding="utf-8")
        doc2 = tmp_path / "real_notes.txt"
        doc2.write_text("Note content", encoding="utf-8")

        docs = [
            ExportDocument(
                id="d1", filename="real_report.pdf", file_path=str(doc1),
                file_size=100, mime_type="application/pdf", file_hash="h1",
                status="processed", created_at=0, updated_at=0,
            ),
            ExportDocument(
                id="d2", filename="real_notes.txt", file_path=str(doc2),
                file_size=50, mime_type="text/plain", file_hash="h2",
                status="uploaded", created_at=0, updated_at=0,
            ),
        ]
        result = export_svc.export_zip(docs, include_originals=True)
        with zipfile.ZipFile(result.path) as zf:
            names = zf.namelist()
            assert "originals/real_report.pdf" in names
            assert "originals/real_notes.txt" in names
            assert zf.read("originals/real_report.pdf").decode("utf-8") == "PDF content"

    def test_zip_missing_originals_skipped(self, export_svc: ExportService) -> None:
        """Missing original files are skipped without error."""
        docs = [
            ExportDocument(
                id="d1", filename="missing.txt", file_path="/nonexistent/missing.txt",
                file_size=100, mime_type="text/plain", file_hash="h1",
                status="uploaded", created_at=0, updated_at=0,
            ),
        ]
        result = export_svc.export_zip(docs, include_originals=True)
        with zipfile.ZipFile(result.path) as zf:
            names = zf.namelist()
            assert "originals/missing.txt" not in names

    def test_zip_empty(self, export_svc: ExportService) -> None:
        result = export_svc.export_zip([])
        assert result.count == 0
        with zipfile.ZipFile(result.path) as zf:
            names = zf.namelist()
            assert len(names) == 4  # all four files even when empty


# =========================================================================
# ExportService — ExportResult
# =========================================================================


class TestExportResult:
    def test_timestamp_auto_set(self) -> None:
        r = ExportResult(path=Path("test.json"), format=ExportFormat.JSON, count=1, size_bytes=100)
        assert r.timestamp > 0

    def test_timestamp_preserved(self) -> None:
        r = ExportResult(path=Path("test.json"), format=ExportFormat.JSON, count=1, size_bytes=100, timestamp=42.0)
        assert r.timestamp == 42.0


# =========================================================================
# ExportService — Edge cases
# =========================================================================


class TestEdgeCases:
    def test_unsupported_format(self, export_svc: ExportService) -> None:
        with pytest.raises(ExportError, match="Unsupported format"):
            export_svc.export_batch(SAMPLE_DOCS, ExportFormat.ZIP)  # ZIP not in direct export

    def test_output_dir_created(self, tmp_path: Path) -> None:
        nested = tmp_path / "a" / "b" / "c"
        svc = ExportService(output_dir=nested)
        assert nested.is_dir()
        result = svc.export_batch(SAMPLE_DOCS[:1], ExportFormat.JSON)
        assert result.path.is_file()

    def test_unicode_filenames(self, export_svc: ExportService) -> None:
        result = export_svc.export_batch([SAMPLE_DOCS[2]], ExportFormat.JSON)
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert data[0]["filename"] == "héllo_世界.txt"

    def test_large_extracted_text(self, export_svc: ExportService) -> None:
        big_text = "Lorem ipsum dolor sit amet. " * 10000
        doc = ExportDocument(
            id="big", filename="big.txt", file_path="/tmp/big.txt",
            file_size=100000, mime_type="text/plain", file_hash="big",
            status="processed", created_at=0, updated_at=0,
            extracted_text=big_text, confidence_score=0.5,
        )
        result = export_svc.export_single(doc, ExportFormat.JSON)
        assert result.count == 1
        data = json.loads(result.path.read_text(encoding="utf-8"))
        assert len(data[0]["extracted_text"]) > 100000


# =========================================================================
# Helper functions
# =========================================================================


class TestHelpers:
    def test_human_size_bytes(self) -> None:
        assert _human_size(500) == "500 B"

    def test_human_size_kb(self) -> None:
        assert _human_size(2048) == "2.0 KB"

    def test_human_size_mb(self) -> None:
        assert _human_size(5 * 1024 * 1024) == "5.0 MB"

    def test_human_size_gb(self) -> None:
        assert _human_size(3 * 1024 * 1024 * 1024) == "3.00 GB"

    def test_format_ts(self) -> None:
        result = _format_ts(0)
        assert "1970" in result
